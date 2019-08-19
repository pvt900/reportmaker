## TransferSheet.py Created by Robert Farmer
## For RX01 - DTO
## Supervisor: Jim Croskrey
## This Program will call the ManualReporter Program to crunch the raw SAP Data.
## Once the Data is Crunched it will compared and transfer any NEW data within the Macro Excel File
## Once this is done, you can open the file and run the Macro to upload any new Untimed Parts.
## Upload takes a variable amount of time based on the number of new parts.
##
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict
from UntimeReportTool import ManualReporter
import collections
import datetime
import time
import os
import os.path
import win32com.client


class TransferSheet:
    def __init__(self):
        '''
        Initializes all Variables for use within the ClassProgram
        '''
        Tk().withdraw() # Removes Tkinter Interface popups
        self.utbook = None #Set a None for Later
        self.macbook = load_workbook(
            filename='UploadUntimed.xlsm', keep_vba=True) #keep_vba=True is important to keep the Macro in the file.
    def run_report_tool(self):
        '''
        This Function calls the Untimed Report Tool and Runs the Report tool
        '''
        #This Function Imports the Report Running Tool and Runs the Functions within.
        #Just Call this Function to run the Tool in its entirety. 
        Report = ManualReporter()
        Report.open_sapfile()
        Report.open_tracker()
        Report.trim_report()
        Report.check_rows()
        Report.CreateReport()
        Report.payables()
        Report.worktracker_scanner()
        Report.part_programs()
        Report.count_programs()
        Report.final_report()
        Report.save_workbook()

    def transfer_report(self):
        '''
        This creates the sheets for the two workbooks, clears the contents of the Upload Document
        Then once cleared it transfers the content of the UT. Report.
        '''

        utsheet = self.utbook.worksheets[2] #Sheet w/ Untimed Report 
        macsheet = self.macbook.worksheets[0] # Macro Workbook Sheets 
        new_mac = self.macbook.create_sheet("Untimed Datas", 1) #New Sheet in the Macro book
        new_mac = self.macbook.worksheets[1] #Selects the New Sheet.

        index_mac = {} # Initializes the Dictionary
        #Estimates Only List is to Exclude the Future Product Parts from The Ready Board.

        for mac_rows in macsheet.iter_rows(): # Creates Key-Value Dictionary for Iteration
            key = (mac_rows[3].value, mac_rows[2].value)
            index_mac[key] = mac_rows # The Row w/ a Key of a Part/Op is added to the Dictionary.
        for ut_rows in utsheet.iter_rows(): # Iterates Through the Rows of the Report
            key = (ut_rows[3].value, ut_rows[2].value) # Part/Op is the Key
            if index_mac.get(key) == None: # if the current Key from Row X doesn't exist
                new_mac.append([cell.value for cell in ut_rows]) # Add it to the Sheet

        std = self.macbook.worksheets[0] # Sets the Old Macro Sheet as std
        self.macbook.remove(std) # Deletes the old Sheet so the New Populated Sheet is the only One

    def save_sheet(self):
        '''
        Saves the Work Book as an .XLSM extension to ensure the Macro survives the transfer.
        '''
        self.macbook.save(filename='UploadUntimed.xlsm')

    def run_macro(self):
        '''
        Runs the Macro from the excel sheet. 
        Current Deprecated. Needs Revision to Work. Doesn't run Macro. 
        Do not Run This function it won't work. Run the Macro from within the Excel Workbook Manually
        '''
        # Launch Excel and Open Wrkbook
        # xl=win32com.client.Dispatch("Excel.Application")
        # xl.Workbooks.Open(r"C:\\Users\\UFJUDFM\Desktop\\Python_Project\\UploadUntimed.xlsm") #opens workbook in readonly mode.

        # Run Macro
        # xl.Application.Run("UploadUntimed.xlsm!Module2.AddNew_SP")

        # Save Document and Quit.
        # xl.Application.Save()
        # xl.Application.Quit()

        # Cleanup the com reference.
        #del xl
        if os.path.exists("UploadUntimed.xlsm"):
            xl = win32com.client.Dispatch('Excel.Application')
            xl.Workbooks.Open(Filename="UploadUntimed.xlsm", ReadOnly=1)
            xl.Application.run("AddNew_SP")
            xl.Application.Quit()
            del xl

##
## This is the Main Section, This is where the Program is Called and Ran. 
## Don't edit this Section unless 100% you know what you're editting.

def main():
    x = TransferSheet()
    print("Script Initialized")
    x.run_report_tool()
    print("Untimed Ran")
    currentReport = ("Untimed Report"+str(datetime.date.today()))
    x.utbook = load_workbook(filename= str(currentReport)+".xlsx")
    print("Book Loaded!")
    x.transfer_report()
    print("Data Transferred!")
    x.save_sheet()
    print("Saved! & Ready For Macro!")
    # Don't Uncomment these Lines. They are for a Deprecated Function Currently.
    # x.run_macro()
    # print("Done!")

if __name__ == "__main__":
    main()
