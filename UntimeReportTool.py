# Made by Robert Farmer
# For DTO at Waterloo Works
# First Report should be the Oldest, Second Report should be the Newest
# First Report is known as the SAP variables while the Second has the WT variables.
# Don't Ask me Why Its a carryover from older iterations.
# If you ask why there are so many Comments it's to prevent the unwant modification of the code and to help
# Aide in the Repair of the Code in the event it is broken by editing something somewhere.
# This was Programmed Using OpenPyXL and the Documentation is here: https://openpyxl.readthedocs.io/en/stable/
# Use the Documentation to make edits and understand the Module's Syntax
##
# To-DO:
##
##
##
# These are the Module Imports don't touch them if you delete them or change them accidently you will break the corresponding code.

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict
import collections
import datetime
import time

# D I S C L A I M E R
# DO NOT EDIT UNLESS NECESSARY


class ManualReporter:
    def __init__(self):
        '''
        Initializes Variables for use within the Class
        Hides the tkinter pop-up when using the file dialog
        '''
        Tk().withdraw()
        self.error_count = 0 # This is a Counting Variable to Count missed Parts. 
        self.sap_file = None
        self.tracker_file = None
        self.wb_sap = None
        self.wb_wt = None
        self.XT = 0  # X's -> T's
        self.TT = 0  # T's -> T's
        self.spoXT = 0
        self.spoTT = 0
        self.NP = 0  # Non Payables (E)
        self.P = 0  # Payables (X)
        self.PSPO = 0 # Payables for SPO(X)
        self.NPSPO = 0 # NonPayables for SPO(E)
        self.total_dto = 0  # Total Untimed Parts for DTO
        self.total_spo = 0 # Total Untimed Parts for SPO
        self.deadrows = []  # Rows to Exclude
        self.changes = set()  # Part Changes
        # Variables for Worktracker Func.
        self.x_issued = 0
        self.t_issued = 0
        self.me_approval = 0
        self.x_issued_spo = 0
        self.t_issued_spo = 0
        self.me_approval_spo = 0
        self.costhr = 0
        # Part Program Variables - Add Program Variables here
        # It is now in a Dictionary Format "Key":Value
        # To add a Program type the EXACT! program name that shows in the report
        # Like so: "F13B":0, then follow with a 0. The program should pickup on it by itself.
        self.dept_spo = ["RX01225", "RX01303", "RX01304", "RX01314", "RX01338"]
        self.dto = {'F9':0,"F9A":0,"F9B":0,"7RMY20":0,"LYNX":0, "Pre-IT4":0,"Saturn":0,"Legacy":0,
        "Tooling":0,"Insourced":0,"Aeros":0,"Isis":0,"None":0,"RCI":0,"Multiple":0,"Maximus":0,"Leopard":0,"F11A":0,"F9X":0,"NGT":0}

        self.spo = {'F9':0,"F9A":0,"F9B":0,"7RMY20":0,"LYNX":0, "Pre-IT4":0,"Saturn":0,"Legacy":0,
        "Tooling":0,"Insourced":0,"Aeros":0,"Isis":0,"None":0,"RCI":0,"Multiple":0,"Maximus":0,"Leopard":0,"F11A":0,"F9X":0,"NGT":0}

    def open_sapfile(self):
        '''
        Sets the sap_file variable to be the file selected in the file dialog. (Should be Last Month's Untimed Report)
        Loads the File Workbook as the variable self.wb_sap
        Optional Bit to Create a Backup of the SAP Report so that if Errors Occur a Fresh Clean Copy is Present
        '''
        self.sap_file = askopenfilename(title="Open Last Ran Untimed Report")
        self.wb_sap = load_workbook(filename=self.sap_file)
        # Code to create a backup File in-case of Error or Fault
        #copyfile = "Untimed_Report_Old_" + str(datetime.date.today())+".xlsx"
        # self.wb_sap.save(copyfile)

    def open_tracker(self):
        '''
        Sets the wb_wt variable as the Current Month's SAP Dump.
        Loads the File Workbook as the variable self.wb_wt
        Optional Bit to Create a Backup of the Second SAP Report so that if Error Occurs a Clean Copy is Present.
        '''
        self.tracker_file = askopenfilename(title="Open Latest SAP Dump")
        self.wb_wt = load_workbook(filename=self.tracker_file)
        #copyfile = "Untimed_Report_New_"+str(datetime.date.today())+".xlsx"
        # self.wb_wt.save(copyfile)

    def trim_report(self):
        '''
        Trims MiscData from Dump. It then Creates the Untimed with Ts/Ps
        '''
        wt = self.wb_wt.worksheets[0]  # Sets the First Sheet in the Second Report as the var wt
        # Create a Spare Sheet in the First Report to place the Adjusted Data
        ws1 = self.wb_wt.create_sheet("Parts Sorted", 1)
        ws1 = self.wb_wt.worksheets[1]  # Sets ws1 as the Active Second Sheet for New Data\

        for row in wt.iter_rows():
            print(row)
            if row[7].value == "MS":
                ws1.append([cell.value for cell in row])
                continue
            if row[4].value != "F" and "AS" not in row[8].value and int(row[7].value) < 60 and row[0].value != "" and "RELEASE IN" not in row[9].value and "TEMPORARY CHECK" not in row[9].value and "ME" not in row[1].value and "72531" not in row[1].value and "76823" not in row[1].value and "INTERFACTORY ROUTING" not in row[9].value and "72515" not in row[1].value and "Release In" not in row[1].value and "Temporary Check" not in row[1].value:
                ws1.append([cell.value for cell in row])

    def check_rows(self):
        '''
       Compares the Unfiltered Untimed Report from Last Month to This Month.
       It also builds a count of the X's and T's in that Report and Keep a List of them.
       It also creates a second list of changed items that suffered a change in the Std/Val categories.
        '''
        self.deadrows = []
        # Sets First Sheet as Active Sheet -  Second Sheet is Trimmed Report from Last Run.
        sap = self.wb_sap.worksheets[1]
        # Sets Second Sheet as Active Sheet - Second Sheet is the Trimmed Report with T's and P's
        wt = self.wb_wt.worksheets[1]
        indexed_wt = {}
        for wtrow in wt.iter_rows():  # Forloop goes through new report marks out Ts Ps
            if wtrow[4].value in ("T", "P"):
                self.deadrows.append(wtrow)
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key] = wtrow
        saprows = list(sap.iter_rows())
        for saprow in saprows:  # For Each Row in the Old Report
            # Sets PartNum & OpCode as Keys
            key = (saprow[3].value, saprow[2].value)
            # For Each Wtrow with Old Report Rows thathave Matching Keys
            wtrow = indexed_wt.get(key, saprow)
            for column in range(12, 16):
                if wtrow[column].value != saprow[column].value:
                    self.changes.add(wtrow)
                    break
            if wtrow[4].value in ("T", "P"):  # New Report Entry is Marked as T/P
                if saprow[4].value == "X":  # Old Report Entry is Marked as X
                    if saprow[0].value in self.dept_spo:
                        self.spoXT +=1
                    else:
                        self.XT += 1  # Increment X->Ts Counts
                else:
                    if saprow[0].value in self.dept_spo:
                        self.spoTT +=1
                    else:
                        self.TT += 1  # Increment T->T Count
        # This Print Counts the Xs->Ts Discounting any Duplicates.
        #print (len([item for item, count in collections.Counter(lst).items() if count > 1]))
 
    def CreateReport(self):
        '''
        Takes the List made in the check_rows function and creates a new sheet
        It the adds all the items from the Previous sheet that are not in the list. (All X/E)
        Also Creates a changes sheet that has any row that suffered a change in it. 
        Final Output will be a sheet with the Correct number of rows, minus one because Row 1 is a Heading Row.
        '''
        wt = self.wb_wt.worksheets[1]
        # Create a Spare Sheet in the First Report to place the Adjusted Data
        ws1 = self.wb_wt.create_sheet("Untimed Parts", 2)
        # Create a Spare Sheet in the  Report to place the Changed Data
        ws2 = self.wb_wt.create_sheet("Changes in Cells M-P", 3)
        # Sets ws2 as Active Third Sheet for Changes Data
        ws2 = self.wb_wt.worksheets[3]
        # Sets ws1 as the Active Second Sheet for Untimed Data
        ws1 = self.wb_wt.worksheets[2]
        for row in wt.iter_rows():
            if row not in self.deadrows:
                ws1.append([cell.value for cell in row])
        for changed in self.changes:
            ws2.append([cell.value for cell in changed])
        # Subtracting 1 from Total Rows because Row 1 is a Header. Thus Exclude it from Count.


    def final_report(self):
        '''
        This Class will Take the data from the other classes and create a Final Sheet to be used for the Reports.
        It will Write the Data needed in the current iteration of the Reports and append it to the Sheet.
        Then it will make the Data appear as a Table in the very last Excel Sheet.
        '''
        ws = self.wb_wt.create_sheet("DTO Report", 4)
        ws = self.wb_wt.worksheets[4]
        wt4 = self.wb_sap.worksheets[4]
        
        for rows in wt4.iter_rows():  # This Will Append the Previous Data from any Other Reports
                ws.append([cell.value for cell in rows])
        refs = ws.max_row  # This Establishes the Length of that Data
        # Deletes the Summations from Prev. Reports
        ws.delete_rows(refs, refs+1)
        currentDT = datetime.datetime.now()  # Current Data
        # Variable Data is the data that get written to the Report
        data = [[currentDT.strftime("%m/%d/%Y"), (self.dto["F9"] + self.dto["F9A"] + self.dto["F9B"] + self.dto["F9X"]),
            self.dto["F11A"], (self.dto["Insourced"] + self.dto["Tooling"] + self.dto["RCI"]), self.dto["7RMY20"],self.dto["LYNX"],self.dto["None"],
            (self.dto["Legacy"]+self.dto["Pre-IT4"]+self.dto["Multiple"]+self.dto["Aeros"]), self.total_dto,
            self.NP,self.P,self.me_approval,self.x_issued,self.t_issued,self.XT, (self.dto["Maximus"]+self.dto["Saturn"]+self.dto["Isis"]),0,0,0, self.costhr ]]
        for row in data:  # Appends the Data Row to the Report
            ws.append(row)
        tablelength = ws.max_row  # Determines the Length of the Table w/ the Data now
        # Creates Dimensions for the Table based on that Length
        tableref = "A1:U" + str(tablelength)
        # Creates the Table with a Display Name and the ref. Ref is the prev. Line
        tab = Table(displayName="DTOTable", ref=tableref)
        style = TableStyleInfo(name="TableStyleDark2", showFirstColumn=True,
                               showLastColumn=True, showRowStripes=True, showColumnStripes=True)
        # Sets Table Style based on the Table Style information parsed in the two lines above.
        tab.tableStyleInfo = style
        ws.add_table(tab)  # Finally Enables the Table to appear
        # Theese Lines Write the Summations for the Xiss, Tiss, X->T
        ws['M'+str(tablelength+1)] = "=SUM(M2:M" + str(tablelength) + ")"
        ws['N'+str(tablelength+1)] = "=SUM(N2:N" + str(tablelength) + ")"
        ws['O'+str(tablelength+1)] = "=SUM(O2:O" + str(tablelength) + ")"

        # Style of Page Below
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        i = 0
        cols = ws.max_column
        #for i in range(cols + 1):
        #   ws[chr(65+i)+str(1)].alignment = Alignment(wrap_text=True)
            #ws['A'+str(i+1)].alignment = Alignment(wrap_text = True)

    def final_report2(self):
        '''
        This Class will Take the data from the other classes and create a Second Report to be used for SPO.
        It will Write the Data needed in the current iteration of the Reports and append it to the Sheet.
        Then it will make the Data appear as a Table in the very last Excel Sheet.
        '''
        ws = self.wb_wt.create_sheet("SPO Report", 5)
        ws = self.wb_wt.worksheets[5]
        wt4 = self.wb_sap.worksheets[5]

        for rows in wt4.iter_rows():  # This Will Append the Previous Data from any Other Reports
                ws.append([cell.value for cell in rows])
        refs = ws.max_row  # This Establishes the Length of that Data
        # Deletes the Summations from Prev. Reports
        ws.delete_rows(refs, refs+1)
        currentDT = datetime.datetime.now()  # Current Data
        # Variable Data is the data that get written to the Report
        #data = [[currentDT.strftime("%m/%d/%Y"), self.fnine, self.feleven, (self.ins + self.tooling + self.rci),
        #         self.sevenr, self.lynx, self.new, (
        #             self.pre+self.saturn+self.maxim+self.legacy+self.aeros+self.isis),
        #         (self.cipp + self.noncipp), self.NP, self.P, self.me_approval, self.x_issued, self.t_issued, self.XT, 0, 0, 0, 0, 0, self.costhr]]
        # Maximus, Saturn, ISIS - NEw Column, Call it FT4/IT4
        data = [[currentDT.strftime("%m/%d/%Y"), (self.spo["F9"] + self.spo["F9A"] + self.spo["F9B"] + self.spo["F9X"]),
            self.spo["F11A"], (self.spo["Insourced"] + self.spo["Tooling"] + self.spo["RCI"]), self.spo["7RMY20"],self.spo["LYNX"],self.spo["None"],
            (self.spo["Legacy"]+self.spo["Pre-IT4"]+self.spo["Multiple"]+self.spo["Aeros"]), 
            self.total_spo,self.NPSPO,self.PSPO,self.me_approval_spo,self.x_issued_spo,
            self.t_issued_spo,self.spoXT,(self.spo["Maximus"]+self.spo["Saturn"]+self.spo["Isis"]),0,0,0,self.costhr]]
        for row in data:  # Appends the Data Row to the Report
            ws.append(row)
        tablelength = ws.max_row  # Determines the Length of the Table w/ the Data now
        # Creates Dimensions for the Table based on that Length
        tableref = "A1:U" + str(tablelength)
        # Creates the Table with a Display Name and the ref. Ref is the prev. Line
        tab = Table(displayName="SPOTable", ref=tableref)
        style = TableStyleInfo(name="TableStyleDark1", showFirstColumn=True,
                               showLastColumn=True, showRowStripes=True, showColumnStripes=True)
        # Sets Table Style based on the Table Style information parsed in the two lines above.
        tab.tableStyleInfo = style
        ws.add_table(tab)  # Finally Enables the Table to appear
        # Theese Lines Write the Summations for the Xiss, Tiss, X->T
        ws['M'+str(tablelength+1)] = "=SUM(M2:M" + str(tablelength) + ")"
        ws['N'+str(tablelength+1)] = "=SUM(N2:N" + str(tablelength) + ")"
        ws['O'+str(tablelength+1)] = "=SUM(O2:O" + str(tablelength) + ")"

        # Style of Page Below
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        i = 0
        cols = ws.max_column
        #for i in range(cols + 1):
        #   ws[chr(65+i)+str(1)].alignment = Alignment(wrap_text=True)
            #ws['A'+str(i+1)].alignment = Alignment(wrap_text = True)

    def payables(self):
        '''
        Counts the Payables & Non-Payables in the Untimed Report.
        '''
        sheet = self.wb_wt.worksheets[2]
        for row in sheet.iter_rows():
            if row[0].value in self.dept_spo:
                if row[4].value is "X":
                    self.PSPO += 1
                elif row[4].value is "E":
                    self.NPSPO += 1
            else:
                if row[4].value is "X":
                    self.P += 1
                elif row[4].value is "E":
                    self.NP += 1

    def save_workbook(self):
        '''
        Saves the Workbook in Current Directory.
        '''
        newfile = "Untimed Report"+str(datetime.date.today())+".xlsx."
        self.wb_wt.save(filename=newfile)
        print("Save Done!")

    def part_programs(self):
        '''
        This Function Takes the Part Program Masterlist and Creates a Dictionary from its Contents.
        From there it iterates through the Rows of the Untimed Report.
        It then appends the Part Program Column and the corresponding Programs
        '''
        wb = load_workbook(
            filename="partprogrammaster.xlsx")  # This Should be the Title of the Part Program Master File.
        wt = self.wb_wt.worksheets[2]
        sheet = wb.active
        indexed_programs = {}
        i = 1
        for row in sheet.iter_rows():
            key = (row[0].value)
            indexed_programs[key] = row[1].value
       #del indexed_programs[None]
        wt.insert_cols(12)
        for rows in wt.iter_rows():
            keys = (rows[3].value)
            wt["L"+str(i)] = str(indexed_programs.get(keys))
            i += 1

    def count_programs(self):
        '''
        Counts Every Part Program in the Untimed Report
        New Programs can be added using the structure below.
        Take the name of the Program as it appears in the Masterlist and
        Surround it with a set of Quotation Marks (" ")
        Then insert similar to any of the statements below.
        Create a new variable up top in the init like so: self.<PartProgramname> = 0
        The variable needs to be created and intialized to 0 for it to work.
        '''
        wt = self.wb_wt.worksheets[2]
        for row in wt.iter_rows():
            if row[0].value in self.dept_spo:
                try:
                    self.spo[row[11].value] += 1
                except KeyError:
                    continue
                self.total_spo += 1
            elif row[0].value not in self.dept_spo:
                try:
                    self.dto[row[11].value] += 1
                except KeyError:
                    continue
                self.total_dto += 1
            
            else:
                self.error_count += 1

    def worktracker_scanner(self):
        '''
        Iterates through the Downloaded Work-Tracker Excel Report, Finds the X's & T's and tracks them
        for the Total X's and T's Issued.
        Also Tracks Entries Marked for ME Approval Just In Case.
        How to use:
        Goto Worktracker using Internet Explorer there will be a tiny box next to the part lists.
        Click the box and uptop near the address bar two boxes will appear called ITEMS and LIST
        Click List and Export to Excel. Open the file that it Downloads and save as wtbook.xlsx in the
        same directory as this file.
        ''' 
        SPO = ["225", "303", "304", "314", "338"]
        filedir = 'wtbook.xlsx'
        wb = load_workbook(filename=filedir)
        sheet = wb.active
        print("If your input is blank it will default to today")
        print("Enter the Dates Year-Month-Day, like so: 2019-01-01\n")
        try:
            lower = str(input("Enter the Lower Date: "))
            if lower == "":
                lower = datetime.datetime.now().strftime('%Y-%m-%d')
            else:
                lower = datetime.datetime.strptime(lower, "%Y-%m-%d")
            upper = str(input("Enter the Upper Date: "))
            if upper == "":
                upper = datetime.datetime.now().strftime("%Y-%m-%d")
            else:
                upper = datetime.datetime.strptime(upper, "%Y-%m-%d")
        except:
            lower = datetime.datetime.now().strftime('%Y-%m-%d')
            upper = datetime.datetime.now().strftime("%Y-%m-%d")
            lower = datetime.datetime.strptime(lower, "%Y-%m-%d")
            upper = datetime.datetime.strptime(upper, "%Y-%m-%d")
        # This loop iterates through the Worktracker file. 
        for row in sheet.iter_rows():
            if "Date of Ch" not in str(row[1].value):
                #date = datetime.datetime.strptime(row[1].value[:10], "%Y-%m-%d")
                date = row[1].value
                # its in the Format of "Year-Month-Day" so, to Scan for Feburary 4th 2020 you'd put "2020-02-04"
                if (lower <= date <= upper):
                    if row[7].value != "Std Type":
                        if row[2].value in SPO:
                            if row[7].value == "T":
                                self.t_issued_spo += 1
                            elif row[7].value == "X":
                                self.x_issued_spo += 1
                            else:
                                self.me_approval_spo +=1
                        else:
                            if row[7].value == "T":
                                self.t_issued += 1
                            elif row[7].value == "X":
                                self.x_issued += 1
                            else:
                                self.me_approval += 1
##
## This is the Previous Main Function. It's Use is for Debugging the Untimed Tool only.

## This is the MAIN this will run the Program. Don't touch it unless adding Functions
## While Deprecated because of the TransferSheet Script this can be used to test modifications
## to the code without using the TransferSheet Script, It's Quicker and Easier to Debug Issues.


#def main():
#    x = ManualReporter()
#    x.open_sapfile()
#    x.open_tracker()
#    progS = time.time()
#    x.trim_report()
#    x.check_rows()
#    x.CreateReport()
#    x.payables()
#    x.worktracker_scanner()
#    x.part_programs()
#    x.count_programs()
#    x.final_report()
#    x.save_workbook()
#    progE = time.time()
#    print("Program Runtime: ", progE - progS)
#if __name__ == "__main__":
#     main()
