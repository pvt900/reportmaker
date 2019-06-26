#Made by Robert Farmer
# For DTO at Waterloo Works
#First Report should be the Oldest, Second Report should be the Newest
#First Report is known as the sap variables while the Second has the WT variables.
#To-DO:
# WRITE GUI (OPTIONAL)

from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict
import datetime
import time
#
# D I S C L A I M E R
#  DO NOT EDIT IF YOU DO THE CONSEQUENCES MAY BE DIRE.
#  ONLY EDIT IF NECESSARY IF REQUIRE HELP CONTACT ME
#  OR TRUST PERSONNEL WHO CAN MAKE THE CHANGES.
class ManualReporter:
    def __init__(self):
        '''
        Initializes Variables for use within the Class
        Hides the tkinter pop-up when using the file dialog
        '''
        Tk().withdraw()
        self.sap_file = None
        self.tracker_file = None
        self.wb_sap = None
        self.wb_wt = None
        self.XT = 0
        self.TT = 0
        self.Error = 0
        self.NP = 0 # Non Payables (E)
        self.P = 0 # Payables (X)
        self.deadrows = []
        self.total = 0

    def open_sapfile(self):
        '''
        Sets the sap_file variable to be the first  directory to the SAP Report based on what the User Selects in the File Dialog
        Sets that directory and the file as the current workbook under the variable self.wb_sap
        Creates a Backup of the SAP Report so that if Errors Occur a Fresh Clean Copy is Present
        '''
        self.sap_file = askopenfilename(title="Open Last Month SAP Report",)
        self.wb_sap = load_workbook(filename=self.sap_file)
        # Code to create a backup File in-case of Error or Fault
        #copyfile = "Untimed_Report_Old_" + str(datetime.date.today())+".xlsx"
        #self.wb_sap.save(copyfile)
        #print(self.sap_file)

    def open_tracker(self):
        '''
        Same as Above, sets self.tracker_file as a filedialog which retrieves the file's directory (User Inputted)
        Loads the File Workbook as self.wb_wt
        Creates a Backup of the Second SAP Report so that if Error Occurs a Clean Copy is Present.
        '''
        self.tracker_file = askopenfilename(title="Current Month's RAW SAP Report")
        self.wb_wt = load_workbook(filename=self.tracker_file)
        #copyfile = "Untimed_Report_New_"+str(datetime.date.today())+".xlsx"
        #self.wb_wt.save(copyfile)
        #print(self.tracker_file)
       
    def trim_report(self):
        '''
        Trims Misc Data from Report to make it more manageable for the comparison
        '''
        #Start = time.time()
        wt = self.wb_wt.worksheets[0]#Sets the First Sheet in the Second Report as the var wt
        ws1 = self.wb_wt.create_sheet("Sheet 1 - Sorted", 1)#Create a Spare Sheet in the First Report to place the Adjusted Data
        ws1 = self.wb_wt.worksheets[1]#Sets ws1 as the Active Second Sheet for New Data
        for row in wt.iter_rows():
            if row[4].value in ("F"):
                self.deadrows.append(row)
            elif "AS" in row[8].value:
                self.deadrows.append(row)
            elif row[7].value != "MS" and int(row[7].value) >= 60:
                    self.deadrows.append(row)
            elif row not in self.deadrows and row[0].value == "":
                self.deadrows.append(row)
            else:
                ws1.append([cell.value for cell in row])
        #End = time.time()
        #print("Trim Time: ", End - Start)
                
    def check_rows(self):
        '''
        Compares the OLD Trimmed Report to the NEW Trimmed Report. Creates a Dictionary (google Dictionary in Python if you must)
        And uses said dictionary to compare matching parts, checks if Status is a T or P and adds to a list of rows called Deadrows
        if it is. 
        '''
        #start = time.time()
        sap = self.wb_sap.worksheets[0] #Sets The First Sheet in the Excel Workbook as the variable sap IN THE FUTURE THIS WILL BE MODIFIED TO BE THE SAME NUMBERED SHEET AS BEFORE
        wt = self.wb_wt.worksheets[1]#Sets the Trimmed Sheet in the Second Report as the var wt
        #This Set of Nested Loops goes through Tests for X's to T's Based on the Report - Report Comparison
        #print(datetime.datetime.today())
        indexed_wt = defaultdict(list)
        for wtrow in wt.iter_rows(): #This For-Loop identifies T's/P's with no matches
            if wtrow[4].value in ("T","P"):
                self.deadrows.append(wtrow)
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key].append(wtrow)
            
        for wtrow in wt.iter_rows():#This For-Loop Finds Matches and Eliminates T's and P's
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key].append(wtrow)
        for saprow in sap.iter_rows():
            key = (saprow[3].value, saprow[2].value)
            for wtrow in indexed_wt[key]:
                if wtrow[4].value in ("T","P"): #New Report Entry is Marked as T/P
                    self.deadrows.append(wtrow)
                    if saprow[4].value is "X": #Old Report Entry is Marked as X
                        self.XT += 1#Increment X->Ts Counts
                    else:
                        self.TT += 1#Increment T->T Count
        #endtime = time.time()
        #print("Comp Timing: ", endtime-start)
    def CreateReport(self):
        '''
        Creates a new Sheet in the Latest Report and iterates through the trimmed list and checks if the Row is in the Deadrows list.
        If it is it skips it otherwise it will add it to the new Sheet.
        Final Output will be a sheet with the Correct number of rows, minus one because Row 1 is a Heading Row.
        '''
        #start = time.time()
        wt = self.wb_wt.worksheets[1]
        ws1 = self.wb_wt.create_sheet("Untimed Parts", 2)#Create a Spare Sheet in the First Report to place the Adjusted Data
        ws1 = self.wb_wt.worksheets[2]#Sets ws1 as the Active Second Sheet for New Data
        for row in wt.iter_rows():
            if row not in self.deadrows:
                ws1.append([cell.value for cell in row])
        #end = time.time()
        #print("Creation Timing: ", end-start)
        self.total = ws1.max_row - 1
    def payables(self):
        '''
        Counts the X's and E's for the Payables & Non-Payables in the Untimed Report.
        '''
        #start = time.time()
        sheet = self.wb_wt.worksheets[2]
        for row in sheet.iter_rows():
            if row[4].value is "X":
                self.P += 1
            elif row[4].value is "E":
                self.NP +=1
        #end = time.time()
        #print("Payables Timing: ", end-start)
        
    def save_workbook(self):
        '''
        Saves the Workbook and prints Payables, Non-Payables and X>T's for Report Tracking. 
        '''
        #start = time.time()
        newfile = "Untimed l "+str(datetime.date.today())+".xlsx."
        self.wb_wt.save(filename = newfile)
        #end = time.time()
        #print("Saving Timing: ", end-start)
        return self.XT,self.P,self.NP,self.total
        print("Save Done!")
        print("Non Payables: ", self.NP)
        print("Payables: ", self.P)
        print("X>T:", self.XT)
                
##progS = time.time()
##x = ManualReporter()
##x.open_sapfile()
##x.open_tracker()
##x.trim_report()
##x.check_rows()
##x.CreateReport()
##x.payables()
##x.save_workbook()
##progE = time.time()
##print("Program Runtime: ", progE - progS)
##
