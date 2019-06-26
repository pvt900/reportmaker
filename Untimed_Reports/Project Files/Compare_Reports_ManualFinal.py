#Made by Robert Farmer
# For DTO at Waterloo Works
#First Report should be the Oldest, Second Report should be the Newest
#First Report is known as the SAP variables while the Second has the WT variables.
#Don't Ask me Why Its a carryover from older iterations. 

#To-DO:
# 

# A COMMENT CAN BE DENOTED BY A #-Symbol. Placing it Before code will Comment it out.
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict
import collections 
import datetime
import time
#
# D I S C L A I M E R
#  DO NOT EDIT IF YOU DO THE CONSEQUENCES MAY BE DIRE.
#  ONLY EDIT IF NECESSARY IF REQUIRE HELP CONTACT ME
#  OR TRUST PERSONNEL WHO CAN MAKE THE CHANGES.
class ManualReporter:
    def __init__(self):
        '''
        Initializes Variables for use within the Class
        Hides the tkinter pop-up when using the file dialog
        '''
        Tk().withdraw()
        self.sap_file = None
        self.tracker_file = None
        self.wb_sap = None
        self.wb_wt = None
        self.XT = 0 #X's -> T's
        self.TT = 0 #T's -> T's
        self.NP = 0 # Non Payables (E)
        self.P = 0 # Payables (X)
        self.total = 0 #Total Untimed Parts
        self.deadrows = []
        self.changes = []

    def open_sapfile(self):
        '''
        Sets the sap_file variable to be the first  directory to the SAP Report based on what the User Selects in the File Dialog
        Sets that directory and the file as the current workbook under the variable self.wb_sap
        Creates a Backup of the SAP Report so that if Errors Occur a Fresh Clean Copy is Present
        '''
        self.sap_file = askopenfilename(title="Open Last Ran Untimed Report")
        self.wb_sap = load_workbook(filename=self.sap_file)
        # Code to create a backup File in-case of Error or Fault
        #copyfile = "Untimed_Report_Old_" + str(datetime.date.today())+".xlsx"
        #self.wb_sap.save(copyfile)
        #print(self.sap_file)

    def open_tracker(self):
        '''
        Same as Above, sets self.tracker_file as a filedialog which retrieves the file's directory (User Inputted)
        Loads the File Workbook as self.wb_wt
        Creates a Backup of the Second SAP Report so that if Error Occurs a Clean Copy is Present.
        '''
        self.tracker_file = askopenfilename(title="Open Latest SAP Dump")
        self.wb_wt = load_workbook(filename=self.tracker_file)
        #copyfile = "Untimed_Report_New_"+str(datetime.date.today())+".xlsx"
        #self.wb_wt.save(copyfile)
        #print(self.tracker_file)
       
    def trim_report(self):
        '''
        Trims Misc Data from Report to make it more manageable for the comparison
        '''
        #Start = time.time()
        wt = self.wb_wt.worksheets[0]#Sets the First Sheet in the Second Report as the var wt
        ws1 = self.wb_wt.create_sheet("Sheet 1 - Sorted", 1)#Create a Spare Sheet in the First Report to place the Adjusted Data
        ws1 = self.wb_wt.worksheets[1]#Sets ws1 as the Active Second Sheet for New Data
        for row in wt.iter_rows():
            if row[4].value in ("F"):
                self.deadrows.append(row)
            elif "AS" in row[8].value:
                self.deadrows.append(row)
            elif row[7].value != "MS" and int(row[7].value) >= 60:
                    self.deadrows.append(row)
            elif row not in self.deadrows and row[0].value == "":
                self.deadrows.append(row)
            else:
                ws1.append([cell.value for cell in row])
        #End = time.time()
        #print("Trim Time: ", End - Start)
                
    def check_rows(self):
        '''
        Compares the OLD Trimmed Report to the NEW Trimmed Report. Creates a Dictionary (google Dictionary in Python if you must)
        And uses said dictionary to compare matching parts, checks if Status is a T or P and adds to a list of rows called Deadrows
        if it is. 
        '''
        self.deadrows = []
        lst = []
        #start = time.time()
        sap = self.wb_sap.worksheets[1] #Sets First Sheet as Active Sheet -  Second Sheet is Trimmed Report from Last Run.
        wt = self.wb_wt.worksheets[1]#Sets Second Sheet as Active Sheet - Second Sheet is the Trimmed Report with T's and P's

        indexed_wt = defaultdict(list)
        for wtrow in wt.iter_rows(): #Forloop goes through new report marks out Ts Ps
            if wtrow[4].value in ("T","P"):
                self.deadrows.append(wtrow)
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key].append(wtrow)
            
        for wtrow in wt.iter_rows():# Creates a Dictionary of the New Report's Trimmed Sheet
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key].append(wtrow)
            
        for saprow in sap.iter_rows():# For Each Row in the Old Report
            key = (saprow[3].value, saprow[2].value) #Sets PartNum & OpCode as Keys

            for wtrow in indexed_wt[key]: #For Each Wtrow with Old Report Rows thathave Matching Keys
                if (wtrow[12].value !=saprow[12].value) or (wtrow[13].value !=saprow[13].value) or (wtrow[14].value != saprow[14].value )or (wtrow[15].value != saprow[15].value):
                    if wtrow not in self.changes:
                        self.changes.append(wtrow)
                if wtrow[4].value in ("T","P"): #New Report Entry is Marked as T/P
                    if saprow[4].value is "X": #Old Report Entry is Marked as X
                        lst.append(wtrow)
                        self.XT += 1#Increment X->Ts Counts
                    else:
                        self.TT += 1#Increment T->T Count

        #This Print Counts the Duplicates in the Xs->Ts
        print (len([item for item, count in collections.Counter(lst).items() if count > 1]))
        #endtime = time.time()
        #print("Comp Timing: ", endtime-start)
    def CreateReport(self):
        '''
        Creates a new Sheet in the Latest Report and iterates through the trimmed list and checks if the Row is in the Deadrows list.
        If it is it skips it otherwise it will add it to the new Sheet.
        Final Output will be a sheet with the Correct number of rows, minus one because Row 1 is a Heading Row.
        '''
        #start = time.time()
        wt = self.wb_wt.worksheets[1]
        ws1 = self.wb_wt.create_sheet("Untimed Parts", 2)#Create a Spare Sheet in the First Report to place the Adjusted Data
        ws2 = self.wb_wt.create_sheet("Changes in Cells M-P", 3)#Create a Spare Sheet in the  Report to place the Changed Data
        ws2 = self.wb_wt.worksheets[3]#Sets ws2 as Active Third Sheet for Changes Data
        ws1 = self.wb_wt.worksheets[2]#Sets ws1 as the Active Second Sheet for Untimed Data
        for row in wt.iter_rows():
            if row not in self.deadrows:
                ws1.append([cell.value for cell in row])
        for changed in self.changes:
            ws2.append([cell.value for cell in changed])
        #end = time.time()
        #print("Creation Timing: ", end-start)
        self.total = (ws1.max_row - 1) # Subtracting 1 from Total Rows because Row 1 is a Header. Thus Exclude it from Count.
    def final_report(self):
        ws = self.wb_wt.create_sheet("Final Report",4)
        ws = self.wb_wt.worksheets[4]
        ws['B1'] = "Date"
        ws['C1'] = "F9 Untimed Parts/Ops"
        ws['D1'] = "Total Untimed Part/Ops"
        ws['E1'] = "Non-Payable Est"
        ws['F1'] = "Payable Est"
        ws['G1'] = "Waiting on ME Approval"
        ws['H1'] = "X's Issued"
        ws['I1'] = "T's Issued"
        ws['J1'] = "X to T"
        
    def payables(self):
        '''
        Counts the X's and E's for the Payables & Non-Payables in the Untimed Report.
        '''
        #start = time.time()
        sheet = self.wb_wt.worksheets[2]
        for row in sheet.iter_rows():
            if row[4].value is "X":
                self.P += 1
            elif row[4].value is "E":
                self.NP +=1
        #end = time.time()
        #print("Payables Timing: ", end-start)
        
    def save_workbook(self):
        '''
        Saves the Workbook and Prints Payables, Non-Payables and X>T's for tracking. 
        '''
        #start = time.time()
        newfile = "Untimed l "+str(datetime.date.today())+".xlsx."
        self.wb_wt.save(filename = newfile)
        #end = time.time()
        #print("Saving Timing: ", end-start)
        print("Save Done!")
        print("----------")
        print("Total Non-Payables: ", self.NP)
        print("Total Payables: ", self.P)
        print("Total X>T:", self.XT)
        print("Total Untimed: ", self.total)
        return (self.NP,self.P,self.XT,self.total)
                  
progS = time.time()
x = ManualReporter()
x.open_sapfile()
x.open_tracker()
x.trim_report()
x.check_rows()
x.CreateReport()
x.payables()
x.final_report()
x.save_workbook()
#input("Press Enter to Close This Window. . . .")
progE = time.time()
print("Program Runtime: ", progE - progS)

