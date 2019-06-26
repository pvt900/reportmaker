## Made by Robert Farmer
## For DTO at Waterloo Works
## First Report should be the Oldest, Second Report should be the Newest
## First Report is known as the SAP variables.
## The Second Report has the WT variables.
## Don't Ask me Why Its a carryover from older iterations.
## If you ask why there are so many Comments it's to prevent
## the unwant modification of the code and to help
## Aide in the Repair of the Code in the event it is broken
## 
##
## A COMMENT CAN BE DENOTED BY A #-Symbol. Placing it Before code will Comment it out.
## These are the Module Imports don't touch them if you delete them or change them accidently you will break the corresponding code.

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict
import collections 
import datetime
import time
##
## D I S C L A I M E R
##  DO NOT EDIT IF YOU DO THE CONSEQUENCES MAY BE DIRE.
##  ONLY EDIT IF NECESSARY IF REQUIRE HELP CONTACT ME
##  OR TRUST PERSONNEL WHO CAN MAKE THE CHANGES.
class ManualReporter:
    def __init__(self):
        '''
        Initializes Variables for use within the Class
        Hides the tkinter pop-up when using the file dialog
        '''
        Tk().withdraw()
        self.sap_file = None
        self.tracker_file = None
        self.wb_sap = None
        self.wb_wt = None
        self.XT = 0 #X's -> T's
        self.TT = 0 #T's -> T's
        self.NP = 0 # Non Payables (E)
        self.P = 0 # Payables (X)
        self.total = 0 #Total Untimed Parts
        self.deadrows = []#Rows to Exclude
        self.changes = [] #Part Changes
        #Variables for Worktracker Func.
        self.x_issued = 0
        self.t_issued = 0
        self.me_approval = 0
        #Part Programs
        self.fnine = 0
        self.sevenr = 0
        self.lynx = 0
        self.pre = 0
        self.saturn = 0
        self.legacy = 0
        self.tooling = 0
        self.ins = 0
        self.aeros = 0
        self.isis = 0
        self.new = 0
        self.rci = 0
        self.mult = 0
        self.maxim = 0
        self.leopard = 0
        self.feleven = 0
        self.ninerx = 0
        
    def open_sapfile(self):
        '''
        Sets the sap_file variable to be the first  directory to the SAP Report based on what the User Selects in the File Dialog
        Sets that directory and the file as the current workbook under the variable self.wb_sap
        Creates a Backup of the SAP Report so that if Errors Occur a Fresh Clean Copy is Present
        '''
        self.sap_file = askopenfilename(title="Open Last Ran Untimed Report")
        self.wb_sap = load_workbook(filename=self.sap_file)
        # Code to create a backup File in-case of Error or Fault
        #copyfile = "Untimed_Report_Old_" + str(datetime.date.today())+".xlsx"
        #self.wb_sap.save(copyfile)
        #print(self.sap_file)

    def open_tracker(self):
        '''
        Same as Above, sets self.tracker_file as a filedialog which retrieves the file's directory (User Inputted)
        Loads the File Workbook as self.wb_wt
        Creates a Backup of the Second SAP Report so that if Error Occurs a Clean Copy is Present.
        '''
        self.tracker_file = askopenfilename(title="Open Latest SAP Dump")
        self.wb_wt = load_workbook(filename=self.tracker_file)
        #copyfile = "Untimed_Report_New_"+str(datetime.date.today())+".xlsx"
        #self.wb_wt.save(copyfile)
        #print(self.tracker_file)
       
    def trim_report(self):
        '''
        Trims Misc Data from Report to make it more manageable for the comparison
        '''
        #Start = time.time()
        wt = self.wb_wt.worksheets[0]#Sets the First Sheet in the Second Report as the var wt
        ws1 = self.wb_wt.create_sheet("Sheet 1 - Sorted", 1)#Create a Spare Sheet in the First Report to place the Adjusted Data
        ws1 = self.wb_wt.worksheets[1]#Sets ws1 as the Active Second Sheet for New Data
        for row in wt.iter_rows():
            if row[4].value in ("F"):
                self.deadrows.append(row)
            elif "AS" in row[8].value:
                self.deadrows.append(row)
            elif row[7].value != "MS" and int(row[7].value) >= 60:
                    self.deadrows.append(row)
            elif row not in self.deadrows and row[0].value == "":
                self.deadrows.append(row)
            else:
                ws1.append([cell.value for cell in row])
        #End = time.time()
        #print("Trim Time: ", End - Start)
                
    def check_rows(self):
        '''
        Compares the OLD Trimmed Report to the NEW Trimmed Report. Creates a Dictionary (google Dictionary in Python if you must)
        And uses said dictionary to compare matching parts, checks if Status is a T or P and adds to a list of rows called Deadrows
        if it is. 
        '''
        self.deadrows = []
        #start = time.time()
        sap = self.wb_sap.worksheets[1] #Sets First Sheet as Active Sheet -  Second Sheet is Trimmed Report from Last Run.
        wt = self.wb_wt.worksheets[1]#Sets Second Sheet as Active Sheet - Second Sheet is the Trimmed Report with T's and P's

        indexed_wt = defaultdict(list)
        for wtrow in wt.iter_rows(): #Forloop goes through new report marks out Ts Ps
            if wtrow[4].value in ("T","P"):
                self.deadrows.append(wtrow)
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key].append(wtrow)
            
        for wtrow in wt.iter_rows():# Creates a Dictionary of the New Report's Trimmed Sheet
            key = (wtrow[3].value, wtrow[2].value)
            indexed_wt[key].append(wtrow)
            
        for saprow in sap.iter_rows():# For Each Row in the Old Report
            key = (saprow[3].value, saprow[2].value) #Sets PartNum & OpCode as Keys

            for wtrow in indexed_wt[key]: #For Each Wtrow with Old Report Rows thathave Matching Keys
                if (wtrow[12].value !=saprow[12].value) or (wtrow[13].value !=saprow[13].value) or (wtrow[14].value != saprow[14].value )or (wtrow[15].value != saprow[15].value):
                    if wtrow not in self.changes:
                        self.changes.append(wtrow)
                if wtrow[4].value in ("T","P"): #New Report Entry is Marked as T/P
                    if saprow[4].value is "X": #Old Report Entry is Marked as X
                        lst.append(wtrow)
                        self.XT += 1#Increment X->Ts Counts
                    else:
                        self.TT += 1#Increment T->T Count

        #This Print Counts the Duplicates in the Xs->Ts
        #print (len([item for item, count in collections.Counter(lst).items() if count > 1]))
        #endtime = time.time()
        #print("Comp Timing: ", endtime-start)
    def CreateReport(self):
        '''
        Creates a new Sheet in the Latest Report and iterates through the trimmed list and checks if the Row is in the Deadrows list.
        If it is it skips it otherwise it will add it to the new Sheet.
        Final Output will be a sheet with the Correct number of rows, minus one because Row 1 is a Heading Row.
        '''
        #start = time.time()
        wt = self.wb_wt.worksheets[1]
        ws1 = self.wb_wt.create_sheet("Untimed Parts", 2)#Create a Spare Sheet in the First Report to place the Adjusted Data
        ws2 = self.wb_wt.create_sheet("Changes in Cells M-P", 3)#Create a Spare Sheet in the  Report to place the Changed Data
        ws2 = self.wb_wt.worksheets[3]#Sets ws2 as Active Third Sheet for Changes Data
        ws1 = self.wb_wt.worksheets[2]#Sets ws1 as the Active Second Sheet for Untimed Data
        for row in wt.iter_rows():
            if row not in self.deadrows:
                ws1.append([cell.value for cell in row])
        for changed in self.changes:
            ws2.append([cell.value for cell in changed])
        #end = time.time()
        #print("Creation Timing: ", end-start)
        self.total = (ws1.max_row - 1) # Subtracting 1 from Total Rows because Row 1 is a Header. Thus Exclude it from Count.

    def final_report(self):
        '''
        This Class will Take the data from the other classes and create a Final Sheet to be used for the Reports.
        It will Write the Data needed in the current iteration of the Reports and append it to the Sheet.
        Then it will make the Data appear as a Table in the Excel Sheet.
        '''
        ws = self.wb_wt.create_sheet("Final Report",4)
        ws = self.wb_wt.worksheets[4]
        wt4 = self.wb_sap.worksheets[4]
        
        for rows in wt4.iter_rows():#This Will Append the Previous Data from any Other Reports
            ws.append([cell.value for cell in rows])
        refs = ws.max_row#This Establishes the Length of that Data 
        ws.delete_rows(refs,refs+1)#Deletes the Summations from Prev. Reports
        currentDT = datetime.datetime.now()#Current Data
        #Variable Data is the data that get written to the Report
        data = [[currentDT.strftime("%Y/%m/%d"),self.fnine, self.feleven, (self.ins + self.tooling + self.rci),
                 self.sevenr, self.lynx,self.new,(self.pre+self.saturn+self.maxim+self.legacy+self.aeros+self.isis),
                 self.total,self.NP,self.P,self.me_approval,self.x_issued,self.t_issued,self.XT]] 
        
        for row in data:#Appends the Data Row to the Report
            ws.append(row)
            
        tablelength = ws.max_row #Determines the Length of the Table w/ the Data now
        tableref = "A1:U" + str(tablelength)#Creates Dimensions for the Table based on that Length
        tab = Table(displayName="Data", ref=tableref)#Creates the Table with a Display Name and the ref. Ref is the prev. Line 
        #Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleDark1", showFirstColumn=True,
                       showLastColumn=True, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style#Sets Table Style based on the Table Style information parsed in the two lines above.
        ws.add_table(tab)#Finally Enables the Table to appear
        #Theese Lines Write the Summations for the Xiss, Tiss, X->T
        ws['M'+str(tablelength+1)] = "=SUM(M2:M" + str(tablelength) +")"
        ws['N'+str(tablelength+1)] = "=SUM(N2:N" + str(tablelength) +")"
        ws['O'+str(tablelength+1)] = "=SUM(O2:O" + str(tablelength) +")"
    def payables(self):
        '''
        Counts the X's and E's for the Payables & Non-Payables in the Untimed Report.
        '''
        #start = time.time()
        sheet = self.wb_wt.worksheets[2]
        for row in sheet.iter_rows():
            if row[4].value is "X":
                self.P += 1
            elif row[4].value is "E":
                self.NP +=1
        #end = time.time()
        #print("Payables Timing: ", end-start)
        
    def save_workbook(self):
        '''
        Saves the Workbook and Prints Payables, Non-Payables and X>T's for tracking. 
        '''
        #start = time.time()
        newfile = "Untimed l "+str(datetime.date.today())+".xlsx."
        self.wb_wt.save(filename = newfile)
        #end = time.time()
        #print("Saving Timing: ", end-start)
        print("Save Done!")
        #Below is Print-Outs for Results.
        #Since Code Generates a Final Report there is no need for them
##        print("----------")
##        print("Total Non-Payables: ", self.NP)
##        print("Total Payables: ", self.P)
##        print("Total X>T:", self.XT)
##        print("Total Untimed: ", self.total)
##        return (self.NP,self.P,self.XT,self.total)

    def part_programs(self):
        '''
        This Function Takes the Part Program Masterlist and Creates a Dictionary from its Contents.
        From there it iterates through the Rows of the Untimed Report and appends the Part Program Column and the corresponding Programs
        to them. 
        '''
        wb = load_workbook(filename = "partprogrammaster.xlsx") #This Should be the Title of the Part Program Master File.
        wt = self.wb_wt.worksheets[2]
        sheet = wb.active
        indexed_programs = {}
        i = 1
        for row in sheet.iter_rows():
            key = (row[0].value)
            indexed_programs[key] = row[1].value
        del indexed_programs[None]
        wt.insert_cols(12)
        for rows in wt.iter_rows():
            keys = (rows[3].value)
            wt["L"+str(i)] = str(indexed_programs.get(keys))
            i+=1
            
    def count_programs(self):
        '''
        This Function Counts the Part Programs on the Untimed for Documentation Purposes.
        New Programs can be added using the structure below. take the name of the Program as it appears in the Masterlist and
        surround it with a set of " " then insert similar to any of the statements below.
        Create a new variable up top in the init like so: self.<PartProgramname> = 0
        '''
        wt = self.wb_wt.worksheets[2]
        for row in wt.iter_rows():
            if "F9" in row[11].value:
                self.fnine +=1
            elif "7RMY20" in row[11].value:
                self.sevenr += 1
            elif "LYNX" in row[11].value:
                self.lynx += 1
            elif "Pre-IT4" in row[11].value:
                self.pre += 1
            elif "Saturn" in row[11].value:
                self.saturn += 1
            elif "Legacy" in row[11].value:
                self.legacy +=1
            elif "Tooling" in row[11].value:
                self.tooling +=1
            elif "Insourced" in row[11].value:
                self.ins += 1
            elif "Aeros" in row[11].value:
                self.aeros += 1
            elif "Isis" in row[11].value:
                self.isis += 1
            elif "RCI" in row[11].value:
                self.rci += 1
            elif "Multiple" in row[11].value:
                self.mult += 1
            elif "Maximus" in row[11].value:
                self.maxim += 1
            elif "Leopard" in row[11].value:
                self.leopard += 1
            elif "F11" in row[11].value:
                self.feleven += 1
            elif "9RX" in row[11].value:
                self.ninerx += 1
            elif "None" in row[11].value:
                self.new +=1
            
    def worktracker_scanner(self):
        '''
        Iterates through the Downloaded Work-Tracker Excel Report, Finds the X's & T's and tracks them
        for the Total X's and T's Issued.
        Also Tracks Entries Marked for ME Approval Just In Case.
        How to use:
        Goto Worktracker using Internet Explorer there will be a tiny box next to the part lists.
        Click the box and uptop near the address bar two boxes will appear called ITEMS and LIST
        Click List and Export to Excel. Open the file that it Downloads and save as wtbook.xlsx in the
        same directory as this file.
        '''
        filedir = 'wtbook.xlsx'
        wb = load_workbook(filename=filedir)
        sheet = wb.active
        
        for row in sheet.iter_rows():#This loop iterates through the Worktracker file. To change the month modify the value below.
            if "2019-04" in  str(row[1].value):  #its in the Format of "Year-Month" so, to Scan for Feburary 2020 you'd put "2020-02"
                if row[7].value != "Std Type":
                    if row[7].value is "T":
                        self.t_issued += 1
                    elif row[7].value is "X":
                        self.x_issued += 1
                    else:
                        self.me_approval += 1

progS = time.time()
x = ManualReporter()
x.open_sapfile()
x.open_tracker()
x.trim_report()
x.check_rows()
x.CreateReport()
x.payables()
x.worktracker_scanner()
x.part_programs()
x.count_programs()
x.final_report()
x.save_workbook()
#input("Press Enter to Close This Window. . . .")
progE = time.time()
print("Program Runtime: ", progE - progS)

