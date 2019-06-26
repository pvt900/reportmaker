from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
#TO DO:
#Talk to Jim abt why the results are so large. Needs a Filter or Comparison to the U.T. Report
class Scanner:

    def __init__(self):
        Tk().withdraw() # Hides TK windows
        self.file = None
        self.wb = None
        self.x_issued = 0
        self.t_issued = 0
        self.me_approval = 0
    def open_file(self):
        self.file =  askopenfilename(title="Open Work Tracker Report") # Opens File Dialog to Grab File Location
        self.wb= load_workbook(filename=self.file) # Takes File Location and Opens as Workbook
        
    def track_issues(self):
        '''
        Iterates through the Downloaded Work-Tracker Excel Report, Finds the X's & T's and tracks them
        for the Total X's and T's Issued.
        Also Tracks Entries Marked for ME Approval Just In Case.
        '''
        sheet = self.wb.active
        
        for row in sheet.iter_rows():
            #print(str(row[1].value))
            #print(row[13].value) # Prints All the Program Values
            if "2019-04" in  str(row[1].value):
                if row[7].value != "Std Type":
                    if row[7].value is "T":
                        self.t_issued += 1
                    elif row[7].value is "X":
                        self.x_issued += 1
                    else:
                        self.me_approval += 1
        #self.report_data()

    def report_data(self):
        #print("T's Issued: " + str(self.t_issued))
        #print("X's Issued: " + str(self.x_issued))
        #print("Things Awaiting ME Approval: " + str(self.me_approval))
        return (self.t_issued, self.x_issued, self.me_approval)
##x = Scanner()
##x.open_file()
##x.track_issues()
##x.report_data
